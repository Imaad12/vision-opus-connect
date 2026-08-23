"""ExcelImporter tests, using small synthetic workbooks generated on the
fly with openpyxl (.xlsx/.xlsm) and xlwt (.xls, dev-only test dependency)
— never a real company document, per the Phase 4 testing brief.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import xlwt

from app.importers.excel_importer import ExcelImporter

BOQ_ROWS = [
    ["Item", "Description", "Trade", "Unit", "Qty", "Rate", "Amount"],
    ["1", "Excavation works", "Civil", "m3", "100", "50.00", "5000.00"],
    ["2", "Blockwork", "Civil", "m2", "200", "75.00", "15000.00"],
]


def _write_xlsx(path: Path, *, hidden_sheet: bool = False) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "BOQ"
    for row in BOQ_ROWS:
        sheet.append(row)

    if hidden_sheet:
        hidden = workbook.create_sheet("Notes")
        hidden.append(["Internal note", "not for client"])
        hidden.sheet_state = "hidden"

    workbook.save(path)


def test_xlsx_import_reads_boq_sheet(tmp_path: Path) -> None:
    path = tmp_path / "quote.xlsx"
    _write_xlsx(path)

    result = ExcelImporter().extract(path)

    assert not result.unsupported
    assert len(result.tables) == 1
    assert result.tables[0].name == "BOQ"
    assert result.tables[0].rows[0] == BOQ_ROWS[0]
    assert result.tables[0].rows[1][1] == "Excavation works"


def test_xlsx_import_flags_hidden_sheets_but_still_reads_them(tmp_path: Path) -> None:
    path = tmp_path / "quote_with_hidden.xlsx"
    _write_xlsx(path, hidden_sheet=True)

    result = ExcelImporter().extract(path)

    sheet_names = [table.name for table in result.tables]
    assert "Notes" in sheet_names
    assert any("hidden" in warning.lower() for warning in result.warnings)


def test_xlsm_import_reads_as_excel_workbook(tmp_path: Path) -> None:
    # openpyxl writes the same underlying zip/XML structure for .xlsm; the
    # importer only cares about the extension routing, exercised here.
    path = tmp_path / "quote.xlsm"
    _write_xlsx(path)

    result = ExcelImporter().extract(path)

    assert not result.unsupported
    assert result.tables[0].rows[1][1] == "Excavation works"


def test_xls_import_reads_legacy_workbook(tmp_path: Path) -> None:
    path = tmp_path / "quote.xls"
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("BOQ")
    for row_index, row in enumerate(BOQ_ROWS):
        for col_index, value in enumerate(row):
            sheet.write(row_index, col_index, value)
    workbook.save(str(path))

    result = ExcelImporter().extract(path)

    assert not result.unsupported
    assert result.tables[0].name == "BOQ"
    assert result.tables[0].rows[0] == BOQ_ROWS[0]
    assert result.tables[0].rows[1][1] == "Excavation works"


def test_corrupt_xlsx_is_reported_as_unsupported_not_crashed(tmp_path: Path) -> None:
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"this is not a real xlsx file")

    result = ExcelImporter().extract(path)

    assert result.unsupported is True
    assert result.unsupported_reason


def test_corrupt_xls_is_reported_as_unsupported_not_crashed(tmp_path: Path) -> None:
    path = tmp_path / "corrupt.xls"
    path.write_bytes(b"not a real xls file either")

    result = ExcelImporter().extract(path)

    assert result.unsupported is True
    assert result.unsupported_reason


def test_empty_workbook_reports_a_warning(tmp_path: Path) -> None:
    path = tmp_path / "empty.xlsx"
    openpyxl.Workbook().save(path)

    result = ExcelImporter().extract(path)

    assert not result.unsupported
    assert result.tables == []
    assert any("no non-empty worksheets" in warning.lower() for warning in result.warnings)
