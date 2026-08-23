"""Excel importer: .xlsx / .xlsm via openpyxl, legacy .xls via xlrd.

Combined into one class since both formats produce the same shape of
result (a set of named worksheets, each a grid of cell text) — the Phase 4
brief explicitly allows combining Excel formats where appropriate. `.xlsb`
is handled separately (`app/importers/xlsb_importer.py`) since it needs a
different library (`pyxlsb`) with a different reading API.

Neither branch loads a workbook it cannot open into memory beyond what the
library itself needs — `read_only=True` on openpyxl streams rows rather
than materializing the whole sheet, which matters for large BOQs.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import xlrd

from app.importers.base import BaseImporter, ExtractedTable, RawExtraction


def _cell_to_str(value: object) -> str:
    return "" if value is None else str(value)


class ExcelImporter(BaseImporter):
    extensions = ("xlsx", "xlsm", "xls")

    def extract(self, path: Path) -> RawExtraction:
        if path.suffix.lower() == ".xls":
            return self._extract_xls(path)
        return self._extract_xlsx(path)

    def _extract_xlsx(self, path: Path) -> RawExtraction:
        warnings: list[str] = []
        tables: list[ExtractedTable] = []
        try:
            workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001 - any open failure is reported, not raised
            return RawExtraction(
                unsupported=True,
                unsupported_reason=f"Could not open workbook (corrupt or password-protected): {exc}",
            )

        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if sheet.sheet_state != "visible":
                    warnings.append(f"Sheet '{sheet_name}' is hidden; included below for review.")
                # `merged_cells` isn't exposed on the read-only worksheet
                # object openpyxl gives us here (deliberately used to avoid
                # loading a large workbook entirely into memory) — merged
                # cells still read fine (only the top-left cell keeps its
                # value), we just can't specifically flag them.
                if getattr(sheet, "merged_cells", None) and sheet.merged_cells.ranges:
                    warnings.append(
                        f"Sheet '{sheet_name}' contains merged cells; only the top-left cell of "
                        "each merged range keeps its value."
                    )
                rows: list[list[str]] = []
                for row in sheet.iter_rows(values_only=True):
                    if row is None or all(cell is None for cell in row):
                        continue
                    rows.append([_cell_to_str(cell) for cell in row])
                if rows:
                    tables.append(ExtractedTable(name=sheet_name, rows=rows))
        finally:
            workbook.close()

        if not tables:
            warnings.append("No non-empty worksheets were found in this workbook.")
        return RawExtraction(tables=tables, warnings=warnings)

    def _extract_xls(self, path: Path) -> RawExtraction:
        warnings: list[str] = []
        tables: list[ExtractedTable] = []
        try:
            workbook = xlrd.open_workbook(str(path))
        except Exception as exc:  # noqa: BLE001
            return RawExtraction(
                unsupported=True,
                unsupported_reason=f"Could not open legacy .xls workbook (corrupt or password-protected): {exc}",
            )

        for sheet in workbook.sheets():
            rows: list[list[str]] = []
            for row_index in range(sheet.nrows):
                row_values = sheet.row_values(row_index)
                if all(value == "" or value is None for value in row_values):
                    continue
                rows.append([_cell_to_str(value) for value in row_values])
            if rows:
                tables.append(ExtractedTable(name=sheet.name, rows=rows))

        if not tables:
            warnings.append("No non-empty worksheets were found in this workbook.")
        return RawExtraction(tables=tables, warnings=warnings)
